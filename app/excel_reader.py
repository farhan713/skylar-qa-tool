"""Read a list of natural-language questions from an Excel .xlsx file.

Expected sheet shape (no strict header — first sheet is read):
    Column A: natural language question (required)
    Column B: expected SQL (optional)

Empty rows are skipped. The first row is treated as a header if column A
contains the strings 'natural_language_query', 'question', or 'query'.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_HINTS = {"natural_language_query", "question", "query", "nl_query", "nl"}


def read_questions(xlsx_path: str | Path) -> list[dict[str, Any]]:
    """Return a list of {id, natural_language_query, expected_sql} dicts.

    Raises FileNotFoundError if the file is missing, ValueError if no
    questions can be extracted.
    """
    p = Path(xlsx_path)
    if not p.exists():
        raise FileNotFoundError(f"questions file not found: {p}")

    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        a = "" if row[0] is None else str(row[0]).strip()
        b = "" if (len(row) < 2 or row[1] is None) else str(row[1]).strip()

        if not a:
            continue

        # First non-empty row may be a header — skip if it looks like one
        if idx == 1 and a.lower() in HEADER_HINTS:
            continue

        rows.append({
            "id": len(rows) + 1,
            "natural_language_query": a,
            "expected_sql": b,
        })

    if not rows:
        raise ValueError(f"no usable rows found in {p}")
    return rows


if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) < 2:
        print("usage: python -m app.excel_reader <path-to-xlsx>")
        sys.exit(1)
    out = read_questions(sys.argv[1])
    print(json.dumps(out, indent=2)[:2000])
    print(f"\n[{len(out)} rows]")
